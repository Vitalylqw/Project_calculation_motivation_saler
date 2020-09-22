import numpy as np
import pandas as pd
import datetime
import openpyxl as xl
import re



class Customers_bils_file():
    """Работа с файлом оперативный учет"""
    def __init__(self, file_name):
        self.file = pd.ExcelFile(file_name)

    def get_sheet_month(self, m, y):
        """Получение из оперативного учета листа нужного месяца"""
        answer = pd.read_excel(self.file, sheet_name=m + ' ' + y)
        answer.drop([0], inplace=True)
        answer = answer.iloc[:, 0:10]
        answer.columns = list(answer.loc[1, :])
        answer.drop([1], inplace=True)
        return answer


class manager_deals_file():
    """Работа с файлом сделки менеджеров"""

    def __init__(self, file_name, manager):
        #         Загрузим наш фал и лист конкретного менеджера
        self.file = pd.read_excel(file_name, sheet_name=manager)
        #     Уберем пустые строки снизу
        self.file = self.file[self.file['Клиент'].notnull()]
        self.file.index = self.file.index + 2

    def check_dupl(self):
        """ Проверка на дублирование счетов"""
        return list(self.file[self.file.duplicated('Счет/доставка')].values)

    def list_for_check(self):
        '''Вернет список строк, который нужно проверить'''
        return self.file[self.file['Дата проверки'].isnull() & self.file['Клиент'].notnull()]

class documents_control_file():
    """Работа с файлом наличия подписанных накладных по сделкам"""
    def __init__(self,file_name):
#         Загрузим наш фйал (лист там один)
        self.file = pd.read_excel(file_name)
#     Удалим строки без счета
        self.file.dropna(axis=0,inplace=True)
#     Создадим счет в ноормальном виде
        self.file['Счет']=[re.search('[1-9]\d* от \d{2}.\d{2}\.\d{4}',x)[0] for x in self.file['Счет на оплату'].values if x==x]
    def check_docs(self,bill):
        bills = self.file.loc[self.file['Счет']==bill.strip(),'Подписан'].values
        if ('Нет' in bills) or (len(bills)==0):
            return False
        else:
            return True

class write_to_exel():
    """Запись данных в рабочий файл менеджера"""
    def __init__(self,name,manager):
        self.name=name
        self.file = xl.load_workbook(name)
        self.ws = self.file[manager]
    # Вносим запись
    def  rec(self,row,rate,date):
        self.ws['M'+str(row)]=rate
        self.ws['O' + str(row)] = date
    # Записываем и закрываем файл
    def clos(self):
        self.file.save(self.name)
        self.file.close()
